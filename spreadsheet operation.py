AIM:
To demonstrate python program to read the data from the spreadsheet and write the data
in to the spreadsheet

ALGORITHM:
1. Start
2. Import the openpyxl module.
3. Load the existing Excel file using openpyxl.load_workbook().
4. Select the worksheet to read from using workbook[sheet_name].
5. Read data from specific rows and columns using loops or direct cell references.
6. Create a new workbook (or load another workbook) for writing using Workbook() or
load_workbook().
7. Select or create a sheet for writing.
8. Write the read data into cells using sheet.cell(row, column).value = data.
9. Save the workbook using workbook.save(filename).
10. End

Required Library
You must install openpyxl if not already installed:

PROGRAM:
import openpyxl
from openpyxl import Workbook

# Step 1: Load the workbook and select the sheet to read
read_wb = openpyxl.load_workbook(&quot;input_data.xlsx&quot;)
read_sheet = read_wb.active

# Step 2: Create a new workbook for writing
write_wb = Workbook()
write_sheet = write_wb.active

# Step 3: Read data from read_sheet and write to write_sheet
for i in range(1, read_sheet.max_row + 1):
for j in range(1, read_sheet.max_column + 1):
# Read from input sheet
cell_value = read_sheet.cell(row=i, column=j).value
# Write to output sheet
write_sheet.cell(row=i, column=j).value = cell_value

# Step 4: Save the new workbook
write_wb.save(&quot;output_data.xlsx&quot;)

print(&quot;Data has been read from &#39;input_data.xlsx&#39; and written to &#39;output_data.xlsx&#39;&quot;)

OUTPUT:
Assume your input_data.xlsx file contains:
Name Age
John 25
Alice 30

After execution, a new file output_data.xlsx is created with the same data.
